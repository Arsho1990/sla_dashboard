from flask import Flask, request, render_template, send_file, redirect, session, jsonify, url_for
import pandas as pd
import oracledb
import io
from datetime import datetime, timedelta, timezone
import functools

# Pakistan Standard Time (UTC+5)
PKT = timezone(timedelta(hours=5))


app = Flask(__name__)
app.secret_key = 'your_super_secret_key_here'
app.permanent_session_lifetime = timedelta(minutes=15)  # Session timeout

def get_connection():
    return oracledb.connect(user="asnafees", password="Gryhhhbbn1", dsn="172.16.0.84:1521/orcl.avanza.pk")

def auto_create_year_column(year, conn=None):
    """Auto-create YEAR_* column if it doesn't exist, based on latest YEAR_* column structure"""
    close_conn = False
    if conn is None:
        conn = get_connection()
        close_conn = True
    
    cursor = conn.cursor()
    try:
        col_name = f'YEAR_{year}'
        
        # Check if column already exists
        cursor.execute(f"""
            SELECT COUNT(*) FROM USER_TAB_COLUMNS 
            WHERE TABLE_NAME = 'SLA' AND COLUMN_NAME = '{col_name}'
        """)
        col_exists = cursor.fetchone()[0] > 0
        
        if not col_exists:
            print(f"🔄 Creating column {col_name}...")
            # Create new NUMBER column with default 0
            try:
                cursor.execute(f"ALTER TABLE SLA ADD {col_name} NUMBER DEFAULT 0")
                conn.commit()
                print(f"✅ Column {col_name} created successfully!")
                return True
            except Exception as e:
                print(f"❌ Error creating column: {e}")
                conn.rollback()
                return False
        else:
            print(f"ℹ️  Column {col_name} already exists")
            return True
    except Exception as e:
        print(f"❌ Error in auto_create_year_column: {e}")
        conn.rollback()
        return False
    finally:
        cursor.close()
        if close_conn:
            conn.close()
    
from datetime import datetime, timedelta, timezone

@app.before_request
def make_session_permanent():
    session.permanent = True
    now = datetime.now(PKT)  # Pakistan Standard Time (UTC+5)
    session_modified = session.get('last_interaction')

    if 'username' in session:
        if session_modified:
            try:
                if not isinstance(session_modified, datetime):
                    session_modified = datetime.fromisoformat(session_modified)

                if (now - session_modified).total_seconds() > 600:
                    username = session['username']
                    login_time = session.get('login_time')
                    if login_time and not isinstance(login_time, datetime):
                        login_time = datetime.fromisoformat(login_time)

                    session.clear()

                    # Save session duration
                    if login_time:
                        duration = now - login_time
                        conn = get_connection()
                        cursor = conn.cursor()
                        cursor.execute("""
                            UPDATE USERS
                            SET LAST_SESSION_DURATION = :duration
                            WHERE USERNAME = :username
                        """, {
                            'duration': str(duration),
                            'username': username
                        })
                        conn.commit()
                        cursor.close()
                        conn.close()

                    return redirect(url_for('login'))
            except Exception as e:
                print("Session error:", str(e))

        # Store updated timestamp
        session['last_interaction'] = now.isoformat()

        
        # ✅ Set login time on successful login
#def handle_successful_login(username):
#    session['username'] = username
#    session['login_time'] = datetime.now(timezone.utc).isoformat()
#    session['last_interaction'] = datetime.now(timezone.utc)
#    # Also store role if needed
def handle_successful_login(username):
    now = datetime.now(PKT)
    session['username'] = username
    session['login_time'] = now.isoformat()

    # Update LOGIN_TIME in DB
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE USERS SET LOGIN_TIME = :login_time WHERE USERNAME = :username
    """, {
        'login_time': now,
        'username': username
    })
    conn.commit()
    cursor.close()
    conn.close()


@app.route('/')
def dashboard():
    if 'username' not in session:
        return redirect('/login')
    
    # ✅ Auto-create missing year columns on dashboard load
    try:
        from datetime import datetime
        current_year = datetime.now().year
        for year in range(2021, current_year + 2):  # Current year + next year
            auto_create_year_column(year)
    except Exception as e:
        print(f"Warning: Could not auto-create year columns: {e}")

    # Ensure DOC_TYPE column exists in SLA
    try:
        _conn = get_connection()
        _cur = _conn.cursor()
        _cur.execute("ALTER TABLE SLA ADD DOC_TYPE VARCHAR2(20)")
        _conn.commit()
        _cur.close()
        _conn.close()
    except Exception:
        try: _conn.rollback()
        except: pass
        try: _cur.close(); _conn.close()
        except: pass

    selected_bank = request.args.get('bank')
    
    # ✅ Fetch login/logout timestamp for current user
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT LOGIN_TIME, LOGOUT_TIME 
        FROM USERS 
        WHERE USERNAME = :username
    """, {'username': session['username']})
    login_logout = cursor.fetchone()
    login_time = login_logout[0] if login_logout else None
    logout_time = login_logout[1] if login_logout else None

    conn = get_connection()

    # Fetch column list from SLA excluding the BLOB (SLA_DOC) to avoid LOB read-after-close errors
    _meta_cur = conn.cursor()
    _meta_cur.execute("""
        SELECT COLUMN_NAME FROM USER_TAB_COLUMNS
        WHERE TABLE_NAME = 'SLA' AND COLUMN_NAME != 'SLA_DOC'
        ORDER BY COLUMN_ID
    """)
    _sla_cols = [r[0] for r in _meta_cur.fetchall()]
    _meta_cur.close()
    _col_list = ', '.join(_sla_cols)
    df = pd.read_sql(f"SELECT {_col_list} FROM SLA", conn)

    # Active / Inactive counts — read from SLA table
    # Any value that is NOT exactly 'Inactive' (including 'complete', NULL, etc.) is treated as Active
    try:
        cur2 = conn.cursor()
        cur2.execute("""
            SELECT
                SUM(CASE WHEN UPPER(NVL(STATUS,'ACTIVE')) != 'INACTIVE' THEN 1 ELSE 0 END),
                SUM(CASE WHEN UPPER(NVL(STATUS,'ACTIVE'))  = 'INACTIVE' THEN 1 ELSE 0 END)
            FROM SLA
        """)
        _row = cur2.fetchone()
        active_count   = int(_row[0] or 0)
        inactive_count = int(_row[1] or 0)
        cur2.close()
    except Exception:
        active_count   = len(df)
        inactive_count = 0

    conn.close()

    # Dynamically detect year columns from the database
    year_cols = sorted([col for col in df.columns if col.startswith('YEAR_')])
    
    # Clean year columns
    for col in year_cols:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                      .str.replace(',', '', regex=False)
                      .str.strip()
                      .replace({'None': '0', 'null': '0', '': '0'})
                      .astype(float)
            )

    df['SLA_CURRENCY'] = df['SLA_CURRENCY'].astype(str).str.upper().str.strip().fillna('UNKNOWN')
    all_banks = df['BANK_NAME'].unique().tolist()

    if selected_bank:
        df = df[df['BANK_NAME'] == selected_bank]
    
    def compute_diff_and_increment(row):
        """Auto-calculate Difference = Current Year - Last Year, and Increment % """
        # Get years from year_cols and convert to actual year numbers
        year_values = []
        for col in year_cols:
            year_num = int(col.replace('YEAR_', ''))
            value = row[col] if row[col] > 0 else 0
            year_values.append((year_num, value))
        
        # Sort by year to find consecutive years
        year_values.sort()
        
        # Find latest two non-zero consecutive years
        diff = 0
        increment_pct = 0
        
        for i in range(len(year_values) - 1, 0, -1):
            curr_year, curr_val = year_values[i]
            prev_year, prev_val = year_values[i-1]
            
            # If values are non-zero, calculate difference and increment
            if curr_val > 0:
                diff = curr_val - prev_val
                increment_pct = (diff / prev_val * 100) if prev_val != 0 else 0
                break
        
        return pd.Series([diff, increment_pct], index=['DIFFERENCE', 'INCREMENT'])

    df[['DIFFERENCE', 'INCREMENT']] = df.apply(compute_diff_and_increment, axis=1)

    def safe_float(val):
        try:
            return float(val)
        except:
            return 0.0

    banks = df['BANK_NAME'].tolist()
    
    # Dynamically create year values lists for all available years
    year_values_dict = {}
    for col in year_cols:
        year_num = col.replace('YEAR_', '')
        year_values_dict[f'y{year_num}'] = [safe_float(v) for v in df[col]]

    # Compute year-wise totals by SLA_CURRENCY
    grouped = df.groupby('SLA_CURRENCY')[year_cols].sum().reset_index()

    currency_totals = grouped.to_dict(orient='records')
    
    # Keep available_years as the full column names for template (e.g., ['YEAR_2023', 'YEAR_2024', 'YEAR_2026'])
    available_years = year_cols

    return render_template("dashboard.html",
                           data=df.to_dict(orient='records'),
                           banks=banks,
                           year_values=year_values_dict,
                           available_years=available_years,
                           all_banks=all_banks,
                           selected_bank=selected_bank,
                           role=session.get('role'),
                           currency_totals=currency_totals,
                           active_count=active_count,
                           inactive_count=inactive_count,
                           login_time=login_time,
                           logout_time=logout_time)
                           
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT PASSWORD, ROLE FROM USERS WHERE USERNAME = :username", {'username': username})
        result = cursor.fetchone()
        cursor.close()
        conn.close()

        if result and result[0] == password:
            session['username'] = username
            session['role'] = result[1]  # store role (admin/viewer)
            handle_successful_login(username)
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Invalid credentials')

    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'username' not in session or session.get('role') != 'admin':
        return redirect('/login')

    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        confirm_password = request.form['confirm_password'].strip()
        role = request.form['role'].strip().lower()

        if not username or not password or not confirm_password or role not in ['admin', 'viewer']:
            return render_template('register.html', error='All fields are required with valid role.')

        if password != confirm_password:
            return render_template('register.html', error='Passwords do not match.')

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM USERS WHERE USERNAME = :username", {'username': username})
        if cursor.fetchone()[0] > 0:
            cursor.close()
            conn.close()
            return render_template('register.html', error='Username already exists.')

        cursor.execute("""
            INSERT INTO USERS (USERNAME, PASSWORD, ROLE, USER_CREATION_DATE)
            VALUES (:username, :password, :role, :creation_date)
        """, {
            'username': username,
            'password': password,
            'role': role,
            'creation_date': datetime.now()
        })
        conn.commit()
        cursor.close()
        conn.close()
        return render_template('register.html', success='User registered successfully.')

    return render_template('register.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form.get('username')
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not username or not old_password or not new_password or not confirm_password:
            return render_template('forgot_password.html', error='All fields are required.')

        if new_password != confirm_password:
            return render_template('forgot_password.html', error='New passwords do not match.')

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT PASSWORD FROM USERS WHERE USERNAME = :username", {'username': username})
        result = cursor.fetchone()

        if result and result[0] == old_password:
            now = datetime.now()
            cursor.execute("""
                UPDATE USERS SET 
                    PASSWORD = :new_password, 
                    LAST_PASSWORD_CHANGE = TO_TIMESTAMP(:now, 'YYYY-MM-DD HH24:MI:SS')
                WHERE USERNAME = :username
            """, {
                'new_password': new_password,
                'now': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'username': username
            })
            conn.commit()
            cursor.close()
            conn.close()
            return render_template('forgot_password.html', success='Password successfully changed.')
        else:
            cursor.close()
            conn.close()
            return render_template('forgot_password.html', error='Old password is incorrect.')

    return render_template('forgot_password.html')
    
@app.route('/logout')
def logout():
    username = session.get('username')
    login_time_str = session.get('login_time')

    if username and login_time_str:
        try:
            login_time = datetime.fromisoformat(login_time_str.replace("Z", "+00:00"))
            logout_time = datetime.now(PKT)
            duration = logout_time - login_time

            # Format duration to HH:MM:SS
            total_seconds = int(duration.total_seconds())
            formatted_duration = f"{total_seconds // 3600:02}:{(total_seconds % 3600) // 60:02}:{total_seconds % 60:02}"

            # Save all session info
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE USERS
                SET 
                    LOGOUT_TIME = :logout_time,
                    LAST_SESSION_DURATION = :duration
                WHERE USERNAME = :username
            """, {
                'logout_time': logout_time,
                'duration': formatted_duration,
                'username': username
            })
            conn.commit()
            cursor.close()
            conn.close()

        except Exception as e:
            print(f"Error saving session data: {e}")

    session.clear()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    """Change password route - admin users only"""
    if 'username' not in session:
        return redirect('/login')
    
    # Only admin users can change their own password
    if session.get('role') != 'admin':
        return render_template('change_password.html', error='Only Admin users can change their password', role=session.get('role'))
    
    if request.method == 'POST':
        username = session.get('username')
        old_password = request.form.get('old_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validations
        if not old_password or not new_password or not confirm_password:
            return render_template('change_password.html', error='All fields are required', role=session.get('role'))
        
        if len(new_password) < 4:
            return render_template('change_password.html', error='New password must be at least 4 characters', role=session.get('role'))
        
        if new_password != confirm_password:
            return render_template('change_password.html', error='Passwords do not match', role=session.get('role'))
        
        if old_password == new_password:
            return render_template('change_password.html', error='New password must be different from old password', role=session.get('role'))
        
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Verify old password
            cursor.execute("SELECT PASSWORD FROM USERS WHERE USERNAME = :username", {'username': username})
            result = cursor.fetchone()
            
            if not result or result[0] != old_password:
                cursor.close()
                conn.close()
                return render_template('change_password.html', error='Old password is incorrect', role=session.get('role'))
            
            # Update password
            cursor.execute(
                "UPDATE USERS SET PASSWORD = :new_password WHERE USERNAME = :username",
                {'new_password': new_password, 'username': username}
            )
            conn.commit()
            cursor.close()
            conn.close()
            
            return render_template('change_password.html', success='✅ Password changed successfully! Please login again', role=session.get('role'))
            
        except Exception as e:
            print(f"Error changing password: {e}")
            return render_template('change_password.html', error=f'Error: {str(e)}', role=session.get('role'))
    
    return render_template('change_password.html', role=session.get('role'))


@app.route("/upload", methods=["POST"])
def upload():
    if 'username' not in session:
        return redirect('/login')
    bank = request.form['bank']
    period = request.form['period']
    doc_type = request.form.get('doc_type', 'Agreement')
    file = request.files['file']
    now = datetime.now(PKT)

    ALLOWED_EXTENSIONS = {
        'pdf', 'txt', 'zip', 'rar',
        'xls', 'xlsx',
        'doc', 'docx'
    }

    if file:
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            return f"File type '.{ext}' not allowed. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}", 400

        content = file.read()
        filename = file.filename
        mimetype = file.mimetype or 'application/octet-stream'

        conn = get_connection()
        cursor = conn.cursor()
        # Ensure DOC_TYPE column exists
        try:
            cursor.execute("ALTER TABLE SLA ADD DOC_TYPE VARCHAR2(20)")
            conn.commit()
        except Exception:
            conn.rollback()
        cursor.execute("""
            UPDATE SLA SET
                SLA_DOC = :blob,
                SLA_DOC_FILENAME = :filename,
                SLA_DOC_TYPE = :type,
                DOC_TYPE = :doc_type,
                UPLOAD_DATE = :upload_date,
                MODIFIED_DATE = :modified_date
            WHERE BANK_NAME = :bank AND PERIOD = :period
        """, {
            'blob': content,
            'filename': filename,
            'type': mimetype,
            'doc_type': doc_type,
            'upload_date': now,
            'modified_date': now,
            'bank': bank,
            'period': period
        })
        conn.commit()
        cursor.close()
        conn.close()
    return redirect("/")

@app.route("/download")
def download_file():
    bank = request.args.get('bank')
    period = request.args.get('period')

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT SLA_DOC, SLA_DOC_FILENAME, SLA_DOC_TYPE
            FROM SLA
            WHERE BANK_NAME = :bank AND PERIOD = :period
        """, {'bank': bank, 'period': period})

        row = cursor.fetchone()
        if row and row[0] is not None:
            # ✅ Read LOB data immediately
            blob_data = row[0].read()
            filename = row[1]
            filetype = row[2]
            # Force download only for archive files; everything else opens in browser tab
            ext = filename.rsplit('.', 1)[-1].lower() if filename and '.' in filename else ''
            force_download = ext in ('zip', 'rar')
            return send_file(io.BytesIO(blob_data), download_name=filename, mimetype=filetype, as_attachment=force_download)

        return "File not found", 404
    except Exception as e:
        return f"Error: {str(e)}", 500
    finally:
        cursor.close()
        conn.close()

@app.route("/delete", methods=["POST"])
def delete_file():
    bank = request.form['bank']
    period = request.form['period']
    now = datetime.now(PKT)

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE SLA SET 
            SLA_DOC = NULL, 
            SLA_DOC_FILENAME = NULL, 
            SLA_DOC_TYPE = NULL,
            DOC_TYPE = NULL,
            MODIFIED_DATE = :modified_date
        WHERE BANK_NAME = :bank AND PERIOD = :period
    """, {'bank': bank, 'period': period, 'modified_date': now})
    conn.commit()
    cursor.close()
    conn.close()
    return redirect("/")
    
@app.route("/update_field", methods=["POST"])
def update_field():
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    data = request.get_json()
    bank = data.get("bank")
    period = data.get("period")
    field = data.get("field")
    value = data.get("value")

    # Whitelist to prevent SQL injection
    SLA_ALLOWED_FIELDS = {
        'PERIOD', 'STATUS', 'DOC_TYPE',
        'YEAR_2021','YEAR_2022','YEAR_2023','YEAR_2024',
        'YEAR_2025','YEAR_2026','YEAR_2027','YEAR_2028'
    }
    if not (field in SLA_ALLOWED_FIELDS or field.startswith('YEAR_')):
        return jsonify({"success": False, "message": "Invalid field"}), 400

    try:
        conn = get_connection()
        
        # If field is a YEAR_* column that doesn't exist, auto-create it
        if field.startswith('YEAR_'):
            year = field.replace('YEAR_', '')
            try:
                year_num = int(year)
                # Try to create the column if it doesn't exist
                auto_create_year_column(year_num, conn)
            except:
                pass
        
        cursor = conn.cursor()
        
        # Update field + modified date
        sql = f"""
            UPDATE SLA 
            SET {field} = :value, MODIFIED_DATE = SYSTIMESTAMP
            WHERE BANK_NAME = :bank AND PERIOD = :period
        """
        cursor.execute(sql, {"value": value, "bank": bank, "period": period})
        
        # Check if row was updated, if not check row count
        rows_updated = cursor.rowcount
        print(f"🔄 Update result: {rows_updated} rows affected")

        # Get fresh timestamp
        cursor.execute("""
            SELECT TO_CHAR(MODIFIED_DATE, 'DD-Mon-YYYY HH:MI:SS AM')
            FROM SLA
            WHERE BANK_NAME = :bank AND PERIOD = :period
        """, {"bank": bank, "period": period})
        timestamp_result = cursor.fetchone()
        new_timestamp = timestamp_result[0] if timestamp_result else datetime.now().strftime('%d-%b-%Y %I:%M:%S %p')

        conn.commit()
        cursor.close()
        
        return jsonify(success=True, last_updated=new_timestamp, rows_updated=rows_updated)

    except Exception as e:
        print(f"❌ Error updating field: {e}")
        try:
            conn.rollback()
        except:
            pass
        return jsonify(success=False, message=str(e))

    finally:
        try:
            cursor.close()
            conn.close()
        except:
            pass

#@app.route("/licenses", methods=["GET", "POST"])
#def licenses():
#    if "username" not in session:
#        return redirect("/login")
#
#    if session.get("role") != "admin":
#        return "Unauthorized access", 403
#
#    conn = get_connection()
#    cursor = conn.cursor()
#
#    # --- Handle RDV License Form ---
#    if request.method == "POST" and request.form.get("form_type") == "rdv":
#        cursor.execute("""
#            MERGE INTO RDV_LICENSES r
#            USING (SELECT :client AS client FROM dual) d
#            ON (r.CLIENT_NAME = d.client)
#            WHEN MATCHED THEN
#                UPDATE SET LICENSE_GUID = :license_guid,
#                           MaxNetworks = :max_networks,
#                           MaxAdapters = :max_adapters,
#                           MT = :mt,
#                           TP = :tp,
#                           TCPIP = :tcpip,
#                           Sybase = :sybase,
#                           WebService = :webservice,
#                           MSMQ = :msmq,
#                           MQSeries = :mqseries,
#                           STATUS = :status,
#                           LAST_UPDATED = CURRENT_TIMESTAMP
#            WHEN NOT MATCHED THEN
#                INSERT (CLIENT_NAME, LICENSE_GUID, MaxNetworks, MaxAdapters, MT, TP, TCPIP, Sybase, WebService, MSMQ, MQSeries, STATUS, LAST_UPDATED)
#                VALUES (:client, :license_guid, :max_networks, :max_adapters, :mt, :tp, :tcpip, :sybase, :webservice, :msmq, :mqseries, :status, CURRENT_TIMESTAMP)
#        """, request.form)
#        conn.commit()
#
#    # --- Handle Nimbus License Form ---
#    if request.method == "POST" and request.form.get("form_type") == "nimbus":
#        cursor.execute("""
#            MERGE INTO NIMBUS_LICENSES n
#            USING (SELECT :client AS client FROM dual) d
#            ON (n.CLIENT_NAME = d.client)
#            WHEN MATCHED THEN
#                UPDATE SET LICENSE_GUID = :license_guid,
#                           MaxTerminals = :max_terminals,
#                           LAST_UPDATED = CURRENT_TIMESTAMP
#            WHEN NOT MATCHED THEN
#                INSERT (CLIENT_NAME, LICENSE_GUID, MaxTerminals, LAST_UPDATED)
#                VALUES (:client, :license_guid, :max_terminals, CURRENT_TIMESTAMP)
#        """, request.form)
#        conn.commit()
#
#    # --- Fetch RDV Data ---
#    cursor.execute("SELECT CLIENT_NAME, LICENSE_GUID, MaxNetworks, MaxAdapters, MT, TP, TCPIP, Sybase, WebService, MSMQ, MQSeries, STATUS, LAST_UPDATED FROM RDV_LICENSES")
#    rdv = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]
#
#    # --- Fetch Nimbus Data ---
#    cursor.execute("SELECT CLIENT_NAME, LICENSE_GUID, MaxTerminals, LAST_UPDATED FROM NIMBUS_LICENSES")
#    nimbus = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]
#
#    cursor.close()
#    conn.close()
#
#    return render_template("licenses.html", rdv=rdv, nimbus=nimbus, clients=clients, selected_client=client_filter, role=session.get("role"))

@app.route("/licenses")
def licenses():
    client_filter = request.args.get("client")

    conn = get_connection()
    cur = conn.cursor()

    # Queries for all three license types
    rdv_query = "SELECT * FROM RDV_LICENSES"
    nimbus_query = "SELECT * FROM NIMBUS_LICENSES"
    unison_query = "SELECT * FROM UNISON_LICENSES"
    log_query = "SELECT * FROM LICENSE_EDIT_LOG"

    params = {}
    if client_filter:
        rdv_query += " WHERE CLIENT_NAME = :client"
        nimbus_query += " WHERE CLIENT_NAME = :client"
        unison_query += " WHERE CLIENT_NAME = :client"
        log_query += " WHERE CLIENT_NAME = :client"
        params = {"client": client_filter}

    # Fetch RDV + Nimbus + Unison rows
    cur.execute(rdv_query, params)
    rdv = [dict(zip([col[0] for col in cur.description], row)) for row in cur.fetchall()]

    cur.execute(nimbus_query, params)
    nimbus = [dict(zip([col[0] for col in cur.description], row)) for row in cur.fetchall()]
    
    cur.execute(unison_query, params)
    unison = [dict(zip([col[0] for col in cur.description], row)) for row in cur.fetchall()]
    
    # Fetch Edit Logs (new)
    cur.execute(log_query + " ORDER BY CHANGE_TIME DESC", params)
    logs = [dict(zip([c[0] for c in cur.description], row)) for row in cur.fetchall()]

    # Always refresh clients list (for dropdown) - now includes UNISON_LICENSES
    cur.execute("""
        SELECT DISTINCT CLIENT_NAME FROM (
            SELECT CLIENT_NAME FROM RDV_LICENSES
            UNION
            SELECT CLIENT_NAME FROM NIMBUS_LICENSES
            UNION
            SELECT CLIENT_NAME FROM UNISON_LICENSES
        ) ORDER BY CLIENT_NAME
    """)
    clients = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    # 🚀 Pass the filter to the template so all forms are prefilled
    return render_template(
        "licenses.html",
        rdv=rdv,
        nimbus=nimbus,
        unison=unison,
        logs=logs,
        clients=clients,
        selected_client=client_filter,
        role=session.get("role")
    )

@app.route("/update_license_field", methods=["POST"])
def update_license_field():
    if session.get("role") != "admin":
        return jsonify({"success": False, "message": "Unauthorized"}), 403

    data = request.json
    table = data["table"]
    client = data["client"]
    field = data["field"]
    new_value = data["value"]
    user = session.get("username")

    conn = get_connection()
    cur = conn.cursor()

    try:
        # fetch old value
        cur.execute(f"SELECT {field} FROM {table} WHERE CLIENT_NAME = :client", {"client": client})
        old_value = cur.fetchone()
        old_value = old_value[0] if old_value else None

        # update field
        cur.execute(f"""
            UPDATE {table}
            SET {field} = :val, LAST_UPDATED = SYSTIMESTAMP
            WHERE CLIENT_NAME = :client
        """, {"val": new_value, "client": client})

        # insert log
        cur.execute("""
            INSERT INTO LICENSE_EDIT_LOG (TABLE_NAME, CLIENT_NAME, FIELD_NAME, OLD_VALUE, NEW_VALUE, CHANGED_BY, CHANGE_TIME)
            VALUES (:table, :client, :field, :old, :new, :user, SYSTIMESTAMP)
        """, {
            "table": table, "client": client, "field": field,
            "old": old_value, "new": new_value, "user": user
        })

        # fetch new last_updated
        cur.execute(f"SELECT TO_CHAR(LAST_UPDATED, 'DD-Mon-YYYY HH:MI:SS AM') FROM {table} WHERE CLIENT_NAME = :client", {"client": client})
        last_updated = cur.fetchone()[0]

        conn.commit()
        return jsonify(success=True, last_updated=last_updated)

    except Exception as e:
        conn.rollback()
        return jsonify(success=False, message=str(e))
    finally:
        cur.close()
        conn.close()

from flask import jsonify, request

# -------------------------
# ADD RDV License (already shared earlier)
# -------------------------

#@app.route("/add_nimbus_license", methods=["POST"])
#def add_nimbus_license():
#    data = request.get_json()
#    try:
#        conn = get_connection()
#        cur = conn.cursor()
#
#        last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#
#        cur.execute("""
#            INSERT INTO NIMBUS_LICENSES
#            (CLIENT_NAME, LICENSE_GUID, MAXTERMINALS, LAST_UPDATED)
#            VALUES (:1,:2,:3,:4)
#            RETURNING ID INTO :5
#        """, [
#            data.get("CLIENT_NAME"),
#            data.get("LICENSE_GUID"),
#            data.get("MAXTERMINALS"),
#            last_updated,
#            cur.var(int)
#        ])
#
#        new_id = cur.getvalue(4)
#        conn.commit()
#        return jsonify({
#            "id": new_id,
#            "CLIENT_NAME": data.get("CLIENT_NAME"),
#            "LICENSE_GUID": data.get("LICENSE_GUID"),
#            "MAXTERMINALS": data.get("MAXTERMINALS"),
#            "LAST_UPDATED": last_updated
#        })
#    except Exception as e:
#        return jsonify({"error": str(e)})
#    finally:
#        cur.close()
#        conn.close()


# -------------------------
# INLINE UPDATE ENDPOINT (RDV + NIMBUS)
# -------------------------

from datetime import datetime

@app.route("/update_license", methods=["POST"])
def update_license():
    from datetime import datetime
    data = request.get_json()
    app.logger.info(f"Received batched update request: {data}")

    id_ = data.get("id")
    table = data.get("table")
    updates = data.get("updates", [])

    allowed_tables = {"RDV_LICENSES", "NIMBUS_LICENSES", "UNISON_LICENSES"}
    if table not in allowed_tables:
        return jsonify({"error": "Invalid table"}), 400

    allowed_fields = {
        "RDV_LICENSES": {"CLIENT_NAME","LICENSE_GUID","MAXNETWORKS","MAXADAPTERS","MTS_MPS","TP",
                         "TCPIP","SYBASE","WEBSERVICE","MSMQ","MQSERIES","STATUS"},
        "NIMBUS_LICENSES": {"CLIENT_NAME","LICENSE_GUID","MAXTERMINALS","STATUS"},
        "UNISON_LICENSES": {"CLIENT_NAME","LICENSE_GUID","MAXAGENTSSEATS","STATUS"}
    }

    conn = get_connection()
    cur = conn.cursor()

    try:
        for upd in updates:
            field, value = upd["field"], upd["value"]

            if field not in allowed_fields[table]:
                return jsonify({"error": f"Invalid field {field}"}), 400

            # 🔹 Fetch old value for audit log
            cur.execute(f"SELECT {field}, CLIENT_NAME FROM {table} WHERE ID = :id", {"id": id_})
            row = cur.fetchone()
            old_value, client_name = row if row else (None, None)

            # 🔹 Use positional binds (:1, :2) instead of named
            sql = f"UPDATE {table} SET {field} = :1, LAST_UPDATED = SYSTIMESTAMP WHERE ID = :2"
            app.logger.info(f"Executing SQL: {sql} with params=({value}, {id_})")
            cur.execute(sql, [value, id_])

            # 🔹 Insert into LICENSE_EDIT_LOG
            cur.execute("""
                INSERT INTO LICENSE_EDIT_LOG
                (TABLE_NAME, CLIENT_NAME, FIELD_NAME, OLD_VALUE, NEW_VALUE, CHANGED_BY, CHANGE_TIME)
                VALUES (:1, :2, :3, :4, :5, :6, SYSTIMESTAMP)
            """, [table, client_name, field, old_value, value, session.get("username", "system")])

        conn.commit()
        return jsonify({
            "status": "ok",
            "LAST_UPDATED": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
    except Exception as e:
        conn.rollback()
        app.logger.error(f"Error updating license: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()




@app.route("/license_history_api/<table>/<client>")
def license_history_api(table, client):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT FIELD_NAME, OLD_VALUE, NEW_VALUE, CHANGED_BY,
               TO_CHAR(CHANGE_TIME, 'DD-Mon-YYYY HH:MI:SS AM') AS CHANGE_TIME
        FROM LICENSE_EDIT_LOG
        WHERE TABLE_NAME = :table AND CLIENT_NAME = :client
        ORDER BY CHANGE_TIME DESC
    """, {"table": table, "client": client})
    logs = [dict(zip([d[0] for d in cur.description], row)) for row in cur.fetchall()]

    cur.close()
    conn.close()
    return jsonify(logs)
 
# --- Add RDV License ---
from flask import jsonify
from datetime import datetime

@app.route("/add_rdv_license", methods=["POST"])
def add_rdv_license():
    
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized: Admins only"}), 403
    
    data = request.get_json()
    print("Received RDV insert:", data)  # debug

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO RDV_LICENSES
            (CLIENT_NAME, LICENSE_GUID, MAXNETWORKS, MAXADAPTERS, MTS_MPS, TP, TCPIP, SYBASE, WEBSERVICE, MSMQ, MQSERIES, STATUS, LAST_UPDATED)
            VALUES (:CLIENT_NAME, :LICENSE_GUID, :MAXNETWORKS, :MAXADAPTERS, :MTS_MPS, :TP, :TCPIP, :SYBASE, :WEBSERVICE, :MSMQ, :MQSERIES, :STATUS, SYSTIMESTAMP)
        """, {
            "CLIENT_NAME": data.get("CLIENT_NAME"),
            "LICENSE_GUID": data.get("LICENSE_GUID"),
            "MAXNETWORKS": data.get("MAXNETWORKS"),
            "MAXADAPTERS": data.get("MAXADAPTERS"),
            "MTS_MPS": data.get("MTS_MPS"),
            "TP": data.get("TP"),
            "TCPIP": data.get("TCPIP"),
            "SYBASE": data.get("SYBASE"),
            "WEBSERVICE": data.get("WEBSERVICE"),
            "MSMQ": data.get("MSMQ"),
            "MQSERIES": data.get("MQSERIES"),
            "STATUS": data.get("STATUS"),
        })

        conn.commit()

        cur.execute("""
            SELECT ID, CLIENT_NAME, LICENSE_GUID, MAXNETWORKS, MAXADAPTERS, MTS_MPS, TP, TCPIP,
                   SYBASE, WEBSERVICE, MSMQ, MQSERIES, STATUS,
                   TO_CHAR(LAST_UPDATED, 'YYYY-MM-DD HH24:MI:SS') AS LAST_UPDATED
            FROM RDV_LICENSES
            WHERE ROWID = (SELECT MAX(ROWID) FROM RDV_LICENSES)
        """)
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        result = dict(zip(cols, row))

    except Exception as e:
        conn.rollback()
        print("Error inserting RDV:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

    return jsonify(result)


@app.route("/add_nimbus_license", methods=["POST"])
def add_nimbus_license():
    
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized: Admins only"}), 403
    
    data = request.get_json()
    print("Received Nimbus insert:", data)  # debug log

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO NIMBUS_LICENSES
            (CLIENT_NAME, LICENSE_GUID, MAXTERMINALS, LAST_UPDATED)
            VALUES (:CLIENT_NAME, :LICENSE_GUID, :MAXTERMINALS, SYSTIMESTAMP)
        """, {
            "CLIENT_NAME": data.get("CLIENT_NAME"),
            "LICENSE_GUID": data.get("LICENSE_GUID"),
            "MAXTERMINALS": data.get("MAXTERMINALS"),
        })

        conn.commit()

        # return the inserted row
        cur.execute("""
            SELECT ID, CLIENT_NAME, LICENSE_GUID, MAXTERMINALS,
                   TO_CHAR(LAST_UPDATED, 'YYYY-MM-DD HH24:MI:SS') AS LAST_UPDATED
            FROM NIMBUS_LICENSES
            WHERE ROWID = (SELECT MAX(ROWID) FROM NIMBUS_LICENSES)
        """)
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        result = dict(zip(cols, row))

    except Exception as e:
        conn.rollback()
        print("Error inserting Nimbus:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

    return jsonify(result)

@app.route("/add_unison_license", methods=["POST"])
def add_unison_license():
    
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized: Admins only"}), 403
    
    data = request.get_json()
    print("Received Unison insert:", data)  # debug log

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO UNISON_LICENSES
            (CLIENT_NAME, LICENSE_GUID, MAXAGENTSSEATS, LAST_UPDATED)
            VALUES (:CLIENT_NAME, :LICENSE_GUID, :MAXAGENTSSEATS, SYSTIMESTAMP)
        """, {
            "CLIENT_NAME": data.get("CLIENT_NAME"),
            "LICENSE_GUID": data.get("LICENSE_GUID"),
            "MAXAGENTSSEATS": data.get("MAXAGENTSSEATS"),
        })

        conn.commit()

        # return the inserted row
        cur.execute("""
            SELECT ID, CLIENT_NAME, LICENSE_GUID, MAXAGENTSSEATS,
                   TO_CHAR(LAST_UPDATED, 'YYYY-MM-DD HH24:MI:SS') AS LAST_UPDATED
            FROM UNISON_LICENSES
            WHERE ROWID = (SELECT MAX(ROWID) FROM UNISON_LICENSES)
        """)
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        result = dict(zip(cols, row))

    except Exception as e:
        conn.rollback()
        print("Error inserting Unison:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

    return jsonify(result)
    
import re
from flask import request, redirect, url_for, jsonify, render_template, session

# ---------- Formula Config Helpers ----------
def ensure_formula_config_table(conn, cur):
    """Create SLA_FORMULA_CONFIG with defaults if it does not exist."""
    cur.execute("SELECT COUNT(*) FROM USER_TABLES WHERE TABLE_NAME = 'SLA_FORMULA_CONFIG'")
    if cur.fetchone()[0] == 0:
        cur.execute("""
            CREATE TABLE SLA_FORMULA_CONFIG (
                FORMULA_KEY      VARCHAR2(50)  PRIMARY KEY,
                FORMULA_LABEL    VARCHAR2(200),
                ADDITIONAL_SIGN  NUMBER(2,0)  DEFAULT  1 NOT NULL,
                SUBTRACTION_SIGN NUMBER(2,0)  DEFAULT -1 NOT NULL,
                UPDATED_BY       VARCHAR2(100),
                UPDATED_AT       TIMESTAMP DEFAULT SYSTIMESTAMP
            )
        """)
        for key, label, a, s in [
            ('increment_with_addition', 'Increment With Addition',          1, -1),
            ('increment_per_excl',      'Increment % (Excl. Additional)',  -1,  1),
            ('increment_per_incl',      'Increment % (Incl. Additional)',   1, -1),
        ]:
            cur.execute(
                "INSERT INTO SLA_FORMULA_CONFIG (FORMULA_KEY,FORMULA_LABEL,ADDITIONAL_SIGN,SUBTRACTION_SIGN) VALUES (:k,:l,:a,:s)",
                {'k': key, 'l': label, 'a': a, 's': s}
            )
        conn.commit()

def get_formula_signs(conn, cur):
    """Return (iwa, excl, incl) sign dicts {'add':int,'sub':int} from SLA_FORMULA_CONFIG."""
    ensure_formula_config_table(conn, cur)
    cur.execute("SELECT FORMULA_KEY, ADDITIONAL_SIGN, SUBTRACTION_SIGN FROM SLA_FORMULA_CONFIG")
    raw = {r[0]: {'add': int(r[1] or 1), 'sub': int(r[2] or -1)} for r in cur.fetchall()}
    return (
        raw.get('increment_with_addition', {'add':  1, 'sub': -1}),
        raw.get('increment_per_excl',      {'add': -1, 'sub':  1}),
        raw.get('increment_per_incl',      {'add':  1, 'sub': -1}),
    )

# ---------- Add SLA (form POST) ----------
@app.route("/add_sla", methods=["POST"])
def add_sla():
    # admin-only
    if session.get("role") != "admin":
        return jsonify({"error": "Not authorized"}), 403

    # reading from an HTML form (application/x-www-form-urlencoded or multipart/form-data)
    data = request.form

    # safe conversions
    def to_float(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    bank = data.get("BANK_NAME")
    currency = data.get("CURRENCY_MODE")
    year = data.get("SLA_YEAR")
    current = to_float(data.get("CURRENT_YEAR_COST", 0))
    last = to_float(data.get("LAST_YEAR_COST", 0))
    additional = to_float(data.get("ADDITIONAL_ITEMS_COST", 0))
    subtraction = to_float(data.get("SUBTRACTION_ITEMS_COST", 0))
    status = data.get("STATUS", "Active")

    conn = get_connection()
    cur = conn.cursor()
    try:
        # Get formula signs and compute increments using formula config
        iwa_s, excl_s, incl_s = get_formula_signs(conn, cur)
        increment = current - last
        increment_with_addition = increment + (iwa_s['add'] * additional) + (iwa_s['sub'] * subtraction)
        if last != 0:
            increment_per = ((current + (excl_s['add'] * additional) + (excl_s['sub'] * subtraction) - last) / last) * 100
            increment_per_with_addition = ((current + (incl_s['add'] * additional) + (incl_s['sub'] * subtraction) - last) / last) * 100
        else:
            increment_per = None
            increment_per_with_addition = None

        cur.execute("""
            INSERT INTO BANK_SLA
            (BANK_NAME, CURRENCY_MODE, SLA_YEAR,
             CURRENT_YEAR_COST, LAST_YEAR_COST, ADDITIONAL_ITEMS_COST, SUBTRACTION_ITEMS_COST,
             SLA_INCREMENT, SLA_INCREMENT_WITH_ADDITION,
             INCREMENT_PER, INCREMENT_PER_WITH_ADDITION, STATUS,
             CREATED_AT, UPDATED_AT)
            VALUES (:BANK_NAME, :CURRENCY_MODE, :SLA_YEAR,
             :CURRENT_YEAR_COST, :LAST_YEAR_COST, :ADDITIONAL_ITEMS_COST, :SUBTRACTION_ITEMS_COST,
             :SLA_INCREMENT, :SLA_INCREMENT_WITH_ADDITION,
             :INCREMENT_PER, :INCREMENT_PER_WITH_ADDITION, :STATUS,
             SYSTIMESTAMP, SYSTIMESTAMP)
        """, {
            "BANK_NAME": bank,
            "CURRENCY_MODE": currency,
            "SLA_YEAR": year,
            "CURRENT_YEAR_COST": current,
            "LAST_YEAR_COST": last,
            "ADDITIONAL_ITEMS_COST": additional,
            "SUBTRACTION_ITEMS_COST": subtraction,
            "SLA_INCREMENT": increment,
            "SLA_INCREMENT_WITH_ADDITION": increment_with_addition,
            "INCREMENT_PER": increment_per,
            "INCREMENT_PER_WITH_ADDITION": increment_per_with_addition,
            "STATUS": status
        })
        conn.commit()
    except Exception as e:
        conn.rollback()
        app.logger.error("Error inserting SLA: %s", e)
        cur.close()
        conn.close()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

    # redirect back to SLA page so grid will refresh
    return redirect(url_for('sla_page'))


# ---------- SLA page (list + filters) ----------
@app.route("/sla")
def sla_page():
    if "username" not in session:
        return redirect("/login")

    bank_filter = request.args.get("bank")
    currency_filter = request.args.get("currency")
    year_filter = request.args.get("year")

    conn = get_connection()
    cur = conn.cursor()

    # Ensure STATUS column exists in BANK_SLA
    try:
        cur.execute("ALTER TABLE BANK_SLA ADD STATUS VARCHAR2(20) DEFAULT 'Active'")
        conn.commit()
    except Exception:
        conn.rollback()

    # build query with optional filters
    base_q = """
      SELECT ID, BANK_NAME, CURRENCY_MODE, SLA_YEAR,
             CURRENT_YEAR_COST, LAST_YEAR_COST, ADDITIONAL_ITEMS_COST, SUBTRACTION_ITEMS_COST,
             SLA_INCREMENT, SLA_INCREMENT_WITH_ADDITION,
             INCREMENT_PER, INCREMENT_PER_WITH_ADDITION, STATUS
      FROM BANK_SLA
    """
    where = []
    params = {}
    if bank_filter:
        where.append("BANK_NAME = :bank")
        params['bank'] = bank_filter
    if currency_filter:
        where.append("CURRENCY_MODE = :currency")
        params['currency'] = currency_filter
    if year_filter:
        where.append("SLA_YEAR = :year")
        try:
            params['year'] = int(year_filter)
        except:
            params['year'] = year_filter

    if where:
        base_q += " WHERE " + " AND ".join(where)

    base_q += " ORDER BY BANK_NAME ASC, SLA_YEAR ASC"

    cur.execute(base_q, params)
    rows = [dict(zip([d[0] for d in cur.description], row)) for row in cur.fetchall()]

    # sanitize numbers and compute fallback percentages if DB null
    def safe_float(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    iwa_signs, excl_signs, incl_signs = get_formula_signs(conn, cur)

    for r in rows:
        r['CURRENT_YEAR_COST'] = safe_float(r.get('CURRENT_YEAR_COST', 0))
        r['LAST_YEAR_COST'] = safe_float(r.get('LAST_YEAR_COST', 0))
        r['ADDITIONAL_ITEMS_COST'] = safe_float(r.get('ADDITIONAL_ITEMS_COST', 0))
        r['SUBTRACTION_ITEMS_COST'] = safe_float(r.get('SUBTRACTION_ITEMS_COST', 0))
        last = r['LAST_YEAR_COST']
        curr = r['CURRENT_YEAR_COST']
        addc = r['ADDITIONAL_ITEMS_COST']
        subtr = r['SUBTRACTION_ITEMS_COST']
        # If DB didn't store SLA_INCREMENT, calculate now
        if r.get('SLA_INCREMENT') in (None,):
            r['SLA_INCREMENT'] = curr - last
        if r.get('SLA_INCREMENT_WITH_ADDITION') in (None,):
            r['SLA_INCREMENT_WITH_ADDITION'] = r['SLA_INCREMENT'] + (iwa_signs['add'] * addc) + (iwa_signs['sub'] * subtr)
        # If percent missing, compute safely using formula config
        if r.get('INCREMENT_PER') in (None,):
            if last != 0:
                r['INCREMENT_PER'] = ((curr + (excl_signs['add'] * addc) + (excl_signs['sub'] * subtr) - last) / last) * 100
            else:
                r['INCREMENT_PER'] = None
        if r.get('INCREMENT_PER_WITH_ADDITION') in (None,):
            if last != 0:
                r['INCREMENT_PER_WITH_ADDITION'] = ((curr + (incl_signs['add'] * addc) + (incl_signs['sub'] * subtr) - last) / last) * 100
            else:
                r['INCREMENT_PER_WITH_ADDITION'] = None

        # build a normalized logo filename so template can load reliably
        name = r.get('BANK_NAME') or ''
        # keep only safe chars and replace separators with underscores
        logo_filename = re.sub(r'[^A-Za-z0-9_.-]', '_', name).lower() + '.png'
        r['LOGO_FILENAME'] = logo_filename

    # populate filter lists from BANK_SLA
    cur.execute("SELECT DISTINCT BANK_NAME FROM BANK_SLA WHERE BANK_NAME IS NOT NULL ORDER BY BANK_NAME")
    banks = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT CURRENCY_MODE FROM BANK_SLA WHERE CURRENCY_MODE IS NOT NULL ORDER BY CURRENCY_MODE")
    currencies = [row[0] for row in cur.fetchall()]
    
    # ---------------------------
    # Year-wise totals by currency
    # ---------------------------
    cur.execute("""
        SELECT CURRENCY_MODE,
               SLA_YEAR,
               SUM(CURRENT_YEAR_COST),
               SUM(LAST_YEAR_COST),
               SUM(ADDITIONAL_ITEMS_COST),
               SUM(NVL(SUBTRACTION_ITEMS_COST, 0)),
               SUM(SLA_INCREMENT),
               SUM(SLA_INCREMENT_WITH_ADDITION)
        FROM BANK_SLA
        GROUP BY CURRENCY_MODE, SLA_YEAR
        ORDER BY CURRENCY_MODE, SLA_YEAR
    """)
    currency_totals = [
        {
            "CURRENCY_MODE": r[0],
            "SLA_YEAR": r[1],
            "CURRENT_YEAR_COST": r[2],
            "LAST_YEAR_COST": r[3],
            "ADDITIONAL_ITEMS_COST": r[4],
            "SUBTRACTION_ITEMS_COST": r[5],
            "SLA_INCREMENT": r[6],
            "SLA_INCREMENT_WITH_ADDITION": r[7]
        }
        for r in cur.fetchall()
    ]

    # Year-wise active / inactive counts
    cur.execute("""
        SELECT SLA_YEAR,
               SUM(CASE WHEN UPPER(NVL(STATUS, 'ACTIVE')) = 'ACTIVE' THEN 1 ELSE 0 END),
               SUM(CASE WHEN UPPER(NVL(STATUS, 'ACTIVE')) = 'INACTIVE' THEN 1 ELSE 0 END)
        FROM BANK_SLA
        GROUP BY SLA_YEAR
        ORDER BY SLA_YEAR
    """)
    year_status_counts = [
        {"SLA_YEAR": r[0], "ACTIVE": int(r[1] or 0), "INACTIVE": int(r[2] or 0)}
        for r in cur.fetchall()
    ]

    cur.close()
    conn.close()

    # years for filter (static list)
    years = [2023, 2024, 2025, 2026, 2027]

    return render_template(
        "sla.html",
        sla=rows,
        banks=banks,
        currencies=currencies,
        years=years,
        role=session.get("role"),
        selected_bank=bank_filter,
        selected_currency=currency_filter,
        selected_year=year_filter,
        currency_totals=currency_totals,
        year_status_counts=year_status_counts
    )
    
@app.route("/update_sla/<int:id>", methods=["POST"])
def update_sla(id):
    if "username" not in session or session.get("role") != "admin":
        return {"success": False, "error": "Unauthorized"}, 403

    data = request.get_json()
    field = data.get("field")
    value = data.get("value")

    allowed_fields = ["CURRENCY_MODE", "SLA_YEAR", "CURRENT_YEAR_COST", "LAST_YEAR_COST", "ADDITIONAL_ITEMS_COST", "SUBTRACTION_ITEMS_COST", "STATUS"]
    if field not in allowed_fields:
        return {"success": False, "error": "Invalid field"}, 400

    conn = get_connection()
    cur = conn.cursor()

    try:
        # Update chosen field
        sql = f"UPDATE BANK_SLA SET {field} = :val, UPDATED_AT = SYSTIMESTAMP WHERE ID = :id"
        cur.execute(sql, {"val": value, "id": id})

        # Recompute increments after update using formula config
        iwa_s, excl_s, incl_s = get_formula_signs(conn, cur)
        cur.execute("""
            UPDATE BANK_SLA
            SET SLA_INCREMENT = CURRENT_YEAR_COST - LAST_YEAR_COST,
                SLA_INCREMENT_WITH_ADDITION = (CURRENT_YEAR_COST - LAST_YEAR_COST)
                    + (:iwa_add * NVL(ADDITIONAL_ITEMS_COST, 0))
                    + (:iwa_sub * NVL(SUBTRACTION_ITEMS_COST, 0)),
                INCREMENT_PER = CASE WHEN LAST_YEAR_COST <> 0
                                     THEN (CURRENT_YEAR_COST
                                          + (:excl_add * NVL(ADDITIONAL_ITEMS_COST, 0))
                                          + (:excl_sub * NVL(SUBTRACTION_ITEMS_COST, 0))
                                          - LAST_YEAR_COST) / LAST_YEAR_COST * 100
                                     ELSE NULL END,
                INCREMENT_PER_WITH_ADDITION = CASE WHEN LAST_YEAR_COST <> 0
                                     THEN (CURRENT_YEAR_COST
                                          + (:incl_add * NVL(ADDITIONAL_ITEMS_COST, 0))
                                          + (:incl_sub * NVL(SUBTRACTION_ITEMS_COST, 0))
                                          - LAST_YEAR_COST) / LAST_YEAR_COST * 100
                                     ELSE NULL END,
                UPDATED_AT = SYSTIMESTAMP
            WHERE ID = :id
        """, {
            'iwa_add': iwa_s['add'], 'iwa_sub': iwa_s['sub'],
            'excl_add': excl_s['add'], 'excl_sub': excl_s['sub'],
            'incl_add': incl_s['add'], 'incl_sub': incl_s['sub'],
            'id': id
        })

        conn.commit()
        return {"success": True}
    except Exception as e:
        conn.rollback()
        return {"success": False, "error": str(e)}
    finally:
        cur.close()
        conn.close()

# ---------- Formula Config API ----------
@app.route('/api/sla-formula-config', methods=['GET'])
def get_formula_config():
    conn = get_connection()
    cur = conn.cursor()
    try:
        ensure_formula_config_table(conn, cur)
        cur.execute("""
            SELECT FORMULA_KEY, FORMULA_LABEL, ADDITIONAL_SIGN, SUBTRACTION_SIGN,
                   UPDATED_BY, UPDATED_AT
            FROM SLA_FORMULA_CONFIG ORDER BY FORMULA_KEY
        """)
        rows = [dict(zip([d[0] for d in cur.description], row)) for row in cur.fetchall()]
        for r in rows:
            r['ADDITIONAL_SIGN']  = int(r['ADDITIONAL_SIGN']  or 1)
            r['SUBTRACTION_SIGN'] = int(r['SUBTRACTION_SIGN'] or -1)
            if r.get('UPDATED_AT'):
                r['UPDATED_AT'] = str(r['UPDATED_AT'])
        return jsonify(rows)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route('/api/sla-formula-config', methods=['POST'])
def save_formula_config():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    conn = get_connection()
    cur = conn.cursor()
#New addtion at 19062026
    def safe_int(value, default):
        """Safely convert to int, falling back to default for None/empty/invalid values."""
        if value is None or value == '':
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default
    #try:
    #    ensure_formula_config_table(conn, cur)
    #    for item in data:
    #        key      = item.get('key')
    #        add_sign = int(item.get('additional_sign',  1))
    #        sub_sign = int(item.get('subtraction_sign', -1))
    #        if add_sign not in (-1, 0, 1) or sub_sign not in (-1, 0, 1):
    #            return jsonify({'error': f'Invalid sign value for {key}'}), 400
    #        cur.execute("""
    #            UPDATE SLA_FORMULA_CONFIG
    #            SET ADDITIONAL_SIGN  = :a,
    #                SUBTRACTION_SIGN = :s,
    #                UPDATED_BY       = :u,
    #                UPDATED_AT       = SYSTIMESTAMP
    #            WHERE FORMULA_KEY = :k
    #        """, {'a': add_sign, 's': sub_sign, 'u': session.get('username'), 'k': key})
    #    conn.commit()
    try:
        ensure_formula_config_table(conn, cur)
        for item in data:
            key      = item.get('key')
            add_sign = safe_int(item.get('additional_sign'),  1)
            sub_sign = safe_int(item.get('subtraction_sign'), -1)
            if add_sign not in (-1, 0, 1) or sub_sign not in (-1, 0, 1):
                return jsonify({'error': f'Invalid sign value for {key}'}), 400
            cur.execute("""
                UPDATE SLA_FORMULA_CONFIG
                SET ADDITIONAL_SIGN  = :a,
                    SUBTRACTION_SIGN = :s,
                    UPDATED_BY       = :u,
                    UPDATED_AT       = SYSTIMESTAMP
                WHERE FORMULA_KEY = :k
            """, {'a': add_sign, 's': sub_sign, 'u': session.get('username'), 'k': key})
        conn.commit()
        # Re-read signs and recalculate ALL BANK_SLA rows
        iwa_s, excl_s, incl_s = get_formula_signs(conn, cur)
        cur.execute("""
            UPDATE BANK_SLA SET
                SLA_INCREMENT = CURRENT_YEAR_COST - LAST_YEAR_COST,
                SLA_INCREMENT_WITH_ADDITION = (CURRENT_YEAR_COST - LAST_YEAR_COST)
                    + (:iwa_add * NVL(ADDITIONAL_ITEMS_COST, 0))
                    + (:iwa_sub * NVL(SUBTRACTION_ITEMS_COST, 0)),
                INCREMENT_PER = CASE WHEN LAST_YEAR_COST <> 0
                    THEN (CURRENT_YEAR_COST
                         + (:excl_add * NVL(ADDITIONAL_ITEMS_COST, 0))
                         + (:excl_sub * NVL(SUBTRACTION_ITEMS_COST, 0))
                         - LAST_YEAR_COST) / LAST_YEAR_COST * 100
                    ELSE NULL END,
                INCREMENT_PER_WITH_ADDITION = CASE WHEN LAST_YEAR_COST <> 0
                    THEN (CURRENT_YEAR_COST
                         + (:incl_add * NVL(ADDITIONAL_ITEMS_COST, 0))
                         + (:incl_sub * NVL(SUBTRACTION_ITEMS_COST, 0))
                         - LAST_YEAR_COST) / LAST_YEAR_COST * 100
                    ELSE NULL END,
                UPDATED_AT = SYSTIMESTAMP
        """, {
            'iwa_add':  iwa_s['add'],  'iwa_sub':  iwa_s['sub'],
            'excl_add': excl_s['add'], 'excl_sub': excl_s['sub'],
            'incl_add': incl_s['add'], 'incl_sub': incl_s['sub']
        })
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()

# API Routes for Dashboard Charts
@app.route('/api/sla-usage-data')
def api_sla_usage_data():
    try:
        selected_bank = request.args.get('bank')
        selected_year = request.args.get('year')
        
        print(f"🔍 API Request: bank={selected_bank}, year={selected_year}")
        
        conn = get_connection()
        df = pd.read_sql("SELECT * FROM SLA", conn)
        conn.close()
        
        # Clean year columns
        year_cols = [col for col in df.columns if col.startswith('YEAR_')]
        for col in year_cols:
            df[col] = (
                df[col].astype(str)
                      .str.replace(',', '', regex=False)
                      .str.strip()
                      .replace({'None': '0', 'null': '0', '': '0', 'nan': '0'})
                      .astype(float)
                      .fillna(0)  # Replace any remaining NaN with 0
            )
        
        df['SLA_CURRENCY'] = df['SLA_CURRENCY'].astype(str).str.upper().str.strip().fillna('UNKNOWN')
        
        print(f"\n📋 Initial df shape: {df.shape}")
        print(f"📋 Initial banks: {df['BANK_NAME'].unique().tolist()}")
        
        if selected_bank:
            print(f"  ✓ Filtering by bank: {selected_bank}")
            df = df[df['BANK_NAME'] == selected_bank]
            print(f"  ✓ After bank filter - df shape: {df.shape}")
        
        # ✅ APPLY YEAR FILTER
        if selected_year:
            current_year_col = f'YEAR_{selected_year}'
            print(f"  ✓ Filtering by year: {selected_year} ({current_year_col})")
            year_cols_to_display = [current_year_col] if current_year_col in year_cols else year_cols
        else:
            print(f"  ✓ No year filter selected - will SUM ALL years")
            year_cols_to_display = year_cols
        
        # Currency Distribution
        print(f"  📊 Before currency dist - df shape: {df.shape}")
        print(f"  📊 Available currencies in df: {df['SLA_CURRENCY'].unique().tolist()}")
        
        if selected_year:
            # If specific year selected, use that year only
            current_year_col = f'YEAR_{selected_year}'
            currency_dist = df.groupby('SLA_CURRENCY')[[current_year_col]].sum().reset_index()
            print(f"  📊 Currency dist for {selected_year}: {currency_dist.to_dict('records')}")
            
            currency_distribution = []
            for _, row in currency_dist.iterrows():
                amount = row[current_year_col] if pd.notna(row[current_year_col]) else 0
                print(f"  💰 {row['SLA_CURRENCY']}: {amount} ({current_year_col})")
                currency_distribution.append({
                    'currency': row['SLA_CURRENCY'],
                    'amount': float(amount)
                })
        else:
            # If All Years selected, sum all year columns
            currency_dist = df.groupby('SLA_CURRENCY')[year_cols].sum().reset_index()
            print(f"  📊 Currency dist (ALL YEARS raw): {currency_dist.to_dict('records')}")
            
            currency_distribution = []
            for _, row in currency_dist.iterrows():
                # Sum all years for this currency
                total_amount = sum([row[col] if pd.notna(row[col]) else 0 for col in year_cols])
                print(f"  💰 {row['SLA_CURRENCY']}: {total_amount} (sum of all years)")
                currency_distribution.append({
                    'currency': row['SLA_CURRENCY'],
                    'amount': float(total_amount)
                })
        
        print(f"  ✅ Final currency_distribution: {currency_distribution}")
        
        # Bank Distribution - SEPARATED BY CURRENCY (USD and PKR)
        # ✅ FIX: When year filter is applied, calculate amount from selected year only
        bank_dist_usd = df[df['SLA_CURRENCY'] == 'USD'].groupby('BANK_NAME')[year_cols].sum().reset_index()
        bank_dist_pkr = df[df['SLA_CURRENCY'] == 'PKR'].groupby('BANK_NAME')[year_cols].sum().reset_index()
        
        bank_distribution_usd = []
        bank_distribution_pkr = []
        
        for _, row in bank_dist_usd.iterrows():
            bank_data = {'bank': row['BANK_NAME'], 'currency': 'USD'}
            
            # Add all years' data to the response
            for col in year_cols:
                year = col.replace('YEAR_', '')
                value = row[col] if pd.notna(row[col]) else 0
                value = 0 if pd.isna(value) else float(value)
                bank_data[f'y{year}'] = value
            
            # ✅ FIX: When year is selected, only calculate amount from that year
            if selected_year:
                bank_data['amount'] = bank_data.get(f'y{selected_year}', 0)
            else:
                # When no year selected, sum all years
                total_amount = 0
                for col in year_cols:
                    year = col.replace('YEAR_', '')
                    total_amount += bank_data.get(f'y{year}', 0)
                bank_data['amount'] = total_amount
            
            bank_distribution_usd.append(bank_data)
        
        for _, row in bank_dist_pkr.iterrows():
            bank_data = {'bank': row['BANK_NAME'], 'currency': 'PKR'}
            
            # Add all years' data to the response
            for col in year_cols:
                year = col.replace('YEAR_', '')
                value = row[col] if pd.notna(row[col]) else 0
                value = 0 if pd.isna(value) else float(value)
                bank_data[f'y{year}'] = value
            
            # ✅ FIX: When year is selected, only calculate amount from that year
            if selected_year:
                bank_data['amount'] = bank_data.get(f'y{selected_year}', 0)
            else:
                # When no year selected, sum all years
                total_amount = 0
                for col in year_cols:
                    year = col.replace('YEAR_', '')
                    total_amount += bank_data.get(f'y{year}', 0)
                bank_data['amount'] = total_amount
            
            bank_distribution_pkr.append(bank_data)
        
        bank_distribution = {'USD': bank_distribution_usd, 'PKR': bank_distribution_pkr}
        
        # Trend Data
        trend_data = []
        for _, row in df.iterrows():
            trend_entry = {'bank': row['BANK_NAME'], 'period': row.get('PERIOD', '')}
            for col in year_cols:
                year = col.replace('YEAR_', '')
                value = row[col] if pd.notna(row[col]) else 0
                value = 0 if pd.isna(value) else float(value)  # Ensure no NaN
                trend_entry[f'y{year}'] = value
            trend_data.append(trend_entry)
        
        # Determine current_year for response
        response_year = selected_year if selected_year else 'All'
        
        print(f"\n✅ API Response Summary:")
        print(f"   - Year filter: {response_year}")
        print(f"   - Currency Distribution entries: {len(currency_distribution)}")
        print(f"   - Bank Distribution USD: {len(bank_distribution_usd)}")
        print(f"   - Bank Distribution PKR: {len(bank_distribution_pkr)}")
        print(f"   - Trend Data entries: {len(trend_data)}")
        
        return jsonify({
            'currency_distribution': currency_distribution,
            'bank_distribution': bank_distribution,
            'trend_data': trend_data,
            'current_year': response_year
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/license-usage-data')
def api_license_usage_data():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        # RDV Status - using raw cursor
        rdv_status = []
        try:
            cursor.execute("SELECT STATUS, COUNT(*) as count FROM RDV_LICENSES GROUP BY STATUS")
            rows = cursor.fetchall()
            rdv_status = [{'STATUS': row[0], 'count': row[1]} for row in rows]
            print(f"✅ RDV Status fetched: {rdv_status}")
        except Exception as e:
            print(f"❌ RDV Query Error: {e}")
            rdv_status = []
        
        # Nimbus Status - NIMBUS_LICENSES has no STATUS column, just return total count
        nimbus_status = []
        try:
            cursor.execute("SELECT COUNT(*) as count FROM NIMBUS_LICENSES")
            count = cursor.fetchone()[0]
            if count > 0:
                nimbus_status = [{'STATUS': 'Active', 'count': count}]
            print(f"✅ Nimbus Status fetched: {nimbus_status}")
        except Exception as e:
            print(f"❌ Nimbus Query Error: {e}")
            nimbus_status = []
        
        # Unison Status - UNISON_LICENSES has no STATUS column, just return total count
        unison_status = []
        try:
            cursor.execute("SELECT COUNT(*) as count FROM UNISON_LICENSES")
            count = cursor.fetchone()[0]
            if count > 0:
                unison_status = [{'STATUS': 'Active', 'count': count}]
            print(f"✅ Unison Status fetched: {unison_status}")
        except Exception as e:
            print(f"❌ Unison Query Error: {e}")
            unison_status = []
        
        # License Type Distribution
        license_type_distribution = []
        try:
            cursor.execute("SELECT COUNT(*) FROM RDV_LICENSES")
            rdv_count = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM NIMBUS_LICENSES")
            nim_count = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM UNISON_LICENSES")
            uni_count = cursor.fetchone()[0]
            
            license_type_distribution = [
                {'type': 'RDV', 'count': rdv_count},
                {'type': 'Nimbus', 'count': nim_count},
                {'type': 'Unison', 'count': uni_count}
            ]
            print(f"✅ License Distribution: {license_type_distribution}")
        except Exception as e:
            print(f"❌ Distribution Query Error: {e}")
            license_type_distribution = []
        
        # Features Usage
        features_usage = {}
        try:
            features = ['MTS_MPS', 'TP', 'TCPIP', 'SYBASE', 'WEBSERVICE', 'MSMQ', 'MQSERIES']
            for feature in features:
                cursor.execute(f"SELECT COUNT(*) FROM RDV_LICENSES WHERE {feature} IS NOT NULL AND {feature} NOT IN ('0', 'N', '')")
                count = cursor.fetchone()[0]
                features_usage[feature] = count
            print(f"✅ Features Usage: {features_usage}")
        except Exception as e:
            print(f"❌ Features Query Error: {e}")
            features_usage = {}
        
        # Recent Changes
        recent_changes = []
        try:
            changes_query = """
                SELECT TABLE_NAME, CLIENT_NAME, FIELD_NAME, OLD_VALUE, NEW_VALUE, CHANGED_BY, CHANGE_TIME 
                FROM LICENSE_EDIT_LOG
                ORDER BY CHANGE_TIME DESC
                FETCH FIRST 10 ROWS ONLY
            """
            cursor.execute(changes_query)
            rows = cursor.fetchall()
            recent_changes = [
                {
                    'TABLE_NAME': row[0],
                    'CLIENT_NAME': row[1],
                    'FIELD_NAME': row[2],
                    'OLD_VALUE': row[3],
                    'NEW_VALUE': row[4],
                    'CHANGED_BY': row[5],
                    'CHANGE_TIME': str(row[6])
                }
                for row in rows
            ]
        except Exception as e:
            print(f"❌ Changes Query Error: {e}")
            recent_changes = []
        
        cursor.close()
        conn.close()
        
        print(f"✅ License API Response: RDV={len(rdv_status)}, Nimbus={len(nimbus_status)}, Unison={len(unison_status)}, Features={len(features_usage)}")
        
        return jsonify({
            'rdv_status': rdv_status,
            'nimbus_status': nimbus_status,
            'unison_status': unison_status,
            'license_type_distribution': license_type_distribution,
            'features_usage': features_usage,
            'recent_changes': recent_changes
        })
    except Exception as e:
        print(f"❌ Error in license endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Export Endpoints
@app.route('/api/export-sla', methods=['POST'])
def export_sla():
    """Export SLA data to Excel or PDF"""
    try:
        data = request.get_json()
        export_format = data.get('format', 'xlsx')  # 'xlsx' or 'pdf'
        columns = data.get('columns', [])
        table_data = data.get('data', [])
        
        if not table_data or not columns:
            return jsonify({'error': 'No data to export'}), 400
        
        if export_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'SLA Data'
            
            # Add headers
            header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for row_idx, row_data in enumerate(table_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            # Set column widths
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
            
            # Save to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            from datetime import datetime
            filename = f"SLA_Data_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                           as_attachment=True, download_name=filename)
        
        elif export_format == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            from io import BytesIO
            from datetime import datetime
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(A4))
            elements = []
            
            # Add title
            styles = getSampleStyleSheet()
            title = Paragraph(f"<b>SLA Data Report</b><br/><font size=10>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>", 
                            styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Create table with optimized column widths
            table_data = [columns] + table_data
            
            # Calculate smart column widths based on content type
            # Use dynamic widths to fit all columns on landscape A4
            num_cols = len(columns)
            page_width = 11 * inch  # Landscape A4 width in points
            margins = 0.5 * inch    # Left and right margins
            available_width = page_width - (2 * margins)
            
            # Smart column width distribution
            col_widths = []
            for i, col_name in enumerate(columns):
                col_name_lower = col_name.lower()
                # Wider columns for long text
                if any(x in col_name_lower for x in ['client', 'service', 'status', 'channel', 'platform']):
                    col_widths.append(available_width * 0.15)  # 15% for text columns
                elif 'year' in col_name_lower:
                    col_widths.append(available_width * 0.08)  # 8% for year columns
                else:
                    col_widths.append(available_width * 0.10)  # 10% default
            
            # Normalize to actual available width
            total_weight = sum(col_widths)
            col_widths = [w * available_width / total_weight for w in col_widths]
            
            # Wrap cell text in Paragraph objects for proper text wrapping
            from reportlab.lib.styles import ParagraphStyle
            wrapped_table_data = []
            cell_style = ParagraphStyle(
                'cellstyle',
                fontSize=9,
                leading=11,
                leftIndent=2,
                rightIndent=2,
                alignment=0,  # LEFT
            )
            header_style = ParagraphStyle(
                'headerstyle',
                fontSize=8,
                leading=10,
                leftIndent=2,
                rightIndent=2,
                alignment=0,  # LEFT
                textColor=colors.whitesmoke,
            )
            
            for row_idx, row in enumerate(table_data):
                wrapped_row = []
                for cell_idx, cell_value in enumerate(row):
                    cell_text = str(cell_value) if cell_value else "-"
                    # Use header style for first row, cell style for data
                    style = header_style if row_idx == 0 else cell_style
                    wrapped_row.append(Paragraph(cell_text, style))
                wrapped_table_data.append(wrapped_row)
            
            table = Table(wrapped_table_data, colWidths=col_widths, repeatRows=1)
            
            # Style table with balanced spacing - readable font + good spacing
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), (75/255, 0/255, 130/255)),  # Purple
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),  # Good spacing
                ('LEFTPADDING', (0, 0), (-1, -1), 6),  # Moderate padding
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),  # Moderate padding
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, (245/255, 245/255, 245/255)]),
                ('LEADING', (0, 0), (-1, -1), 12),  # Good line spacing
            ]))
            
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            
            filename = f"SLA_Data_{datetime.now().strftime('%Y-%m-%d')}.pdf"
            return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)
    
    except Exception as e:
        print(f"❌ Export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-licenses', methods=['POST'])
def export_licenses():
    """Export License data to Excel or PDF"""
    try:
        data = request.get_json()
        export_format = data.get('format', 'xlsx')
        table_type = data.get('table_type', 'rdv')  # 'rdv', 'nimbus', or 'unison'
        columns = data.get('columns', [])
        table_data = data.get('data', [])
        
        if not table_data or not columns:
            return jsonify({'error': 'No data to export'}), 400
        
        # Determine table name based on table_type
        if table_type == 'rdv':
            table_name = 'RDV Licenses'
        elif table_type == 'nimbus':
            table_name = 'Nimbus Licenses'
        elif table_type == 'unison':
            table_name = 'Unison Licenses'
        else:
            table_name = 'Licenses'
        
        if export_format == 'xlsx':
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = table_name
            
            # Add headers
            header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for row_idx, row_data in enumerate(table_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            # Set column widths
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
            
            # Save to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            from datetime import datetime
            filename = f"{table_name.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=filename)
        
        elif export_format == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            from io import BytesIO
            from datetime import datetime
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(A4))
            elements = []
            
            # Add title
            styles = getSampleStyleSheet()
            title = Paragraph(f"<b>{table_name} Report</b><br/><font size=10>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>",
                            styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Create table with optimized column widths
            table_data_with_headers = [columns] + table_data
            
            # Calculate smart column widths based on content type
            num_cols = len(columns)
            page_width = 11 * inch  # Landscape A4 width in points
            margins = 0.5 * inch    # Left and right margins
            available_width = page_width - (2 * margins)
            
            # Smart column width distribution
            col_widths = []
            for i, col_name in enumerate(columns):
                col_name_lower = col_name.lower()
                # Wider columns for GUIDs, timestamps, and long text
                if 'guid' in col_name_lower or 'license' in col_name_lower:
                    col_widths.append(available_width * 0.20)  # 20% width for GUIDs
                elif 'updated' in col_name_lower or 'actions' in col_name_lower or 'client' in col_name_lower:
                    col_widths.append(available_width * 0.15)  # 15% for medium columns
                else:
                    col_widths.append(available_width * 0.08)  # 8% for numeric/short columns
            
            # Normalize to actual available width
            total_weight = sum(col_widths)
            col_widths = [w * available_width / total_weight for w in col_widths]
            
            # Wrap cell text in Paragraph objects for proper text wrapping
            from reportlab.lib.styles import ParagraphStyle
            wrapped_table_data = []
            cell_style = ParagraphStyle(
                'cellstyle',
                fontSize=9,
                leading=11,
                leftIndent=2,
                rightIndent=2,
                alignment=0,  # LEFT
            )
            header_style = ParagraphStyle(
                'headerstyle',
                fontSize=8,
                leading=10,
                leftIndent=2,
                rightIndent=2,
                alignment=0,  # LEFT
                textColor=colors.whitesmoke,
            )
            
            for row_idx, row in enumerate(table_data_with_headers):
                wrapped_row = []
                for cell_idx, cell_value in enumerate(row):
                    cell_text = str(cell_value) if cell_value else "-"
                    # Use header style for first row, cell style for data
                    style = header_style if row_idx == 0 else cell_style
                    wrapped_row.append(Paragraph(cell_text, style))
                wrapped_table_data.append(wrapped_row)
            
            table = Table(wrapped_table_data, colWidths=col_widths, repeatRows=1)
            
            # Style table with balanced spacing - readable font + good spacing
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), (75/255, 0/255, 130/255)),  # Purple
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),  # Good spacing
                ('LEFTPADDING', (0, 0), (-1, -1), 6),  # Moderate padding
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),  # Moderate padding
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, (245/255, 245/255, 245/255)]),
                ('LEADING', (0, 0), (-1, -1), 12),  # Good line spacing
            ]))
            
            elements.append(table)
            doc.build(elements)
            output.seek(0)
            
            filename = f"{table_name.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
            return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)
    
    except Exception as e:
        print(f"❌ Export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == "__main__":
    #app.run(debug=True)
    app.run(host='0.0.0.0', port=5000, debug=True)
